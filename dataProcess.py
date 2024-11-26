import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

def process_data():
    # Read the data from Excel files
    node_details_df = pd.read_excel('node_details.xlsx', header=0)
    customer_details_df = pd.read_excel('customer_details.xlsx', header=0)

    # Perform the left join
    result_df = pd.merge(node_details_df, customer_details_df, left_on='FacilityId', right_on='facility_id', how='left')
    result_df = result_df.drop(columns=['facility_id'])

    # Add columns for various device types
    device_types = {
        'OWM7111': 'OWM7111-GDNT-9-dev',
        'OWM0131': 'OWM0131-GDNT-R',
        'FXA5': 'FXA5000-GCNT-K-dev',
        'ST898ZB': 'ST898ZB',
        '3157100': '3157100',
        'ThermoStats': ['3157100', 'ST898ZB'],
        'Centralized Door Sensor': '3323-G',
        'Centralize Water Sensor': '3315-G',
        'Centralize Temp & Humidity Sensor': '3310-G',
        'Centralize Motion Sensor': '3328-G',
        'R-Pi': 'linuxevb',
        'Common Area Camera': ['ggdemo_camera', 'common_area_camera']
    }

    for col, value in device_types.items():
        if isinstance(value, list):
            result_df[col] = result_df['oemModel'].apply(lambda x: 1 if x in value else 0)
        else:
            result_df[col] = result_df['oemModel'].apply(lambda x: 1 if x == value else 0)

    # Group by 'facility_name' and count the devices
    facility_counts = result_df.groupby(['customer_name', 'facility_name']).sum().reset_index()

    # Group by 'customer_name'
    grouped = facility_counts.groupby('customer_name')

    # Create a dictionary of DataFrames, one for each customer
    customer_dfs = {customer_name: df for customer_name, df in grouped}
    column_order = [
        'Customer Name','Facility Name','OWM7111','OWM0131','FXA5','Common Area Camera','ST898ZB','3157100','ThermoStats','Centralize Temp & Humidity Sensor'
                                                                        ,'Centralize Water Sensor','Centralized Door Sensor',
                                                                        'Centralize Motion Sensor','R-Pi','Total'
    ]
    # Add total row to each DataFrame
    for customer_name, df in customer_dfs.items():
        df['Total'] = df[['OWM7111','OWM0131','FXA5','Common Area Camera','ST898ZB','3157100','ThermoStats','Centralize Temp & Humidity Sensor'
                                                                        ,'Centralize Water Sensor','Centralized Door Sensor',
                                                                        'Centralize Motion Sensor','R-Pi']].sum(axis=1)
        df.drop(columns=['DSN','DeviceName','oemModel','DeviceType','FacilityId','lastStatus','customer_id'], inplace=True)
        df.rename(columns={'customer_name':'Customer Name','facility_name':'Facility Name'}, inplace=True)
        df = df[column_order]
        df.drop(columns=['ThermoStats'], inplace=True)
        total_values = df.iloc[:, 2:].sum()
        total_values['Facility Name'] = 'Total'
        df.loc[max(df.index)+1] = total_values
        # df = df.append(total_values, ignore_index=True)
        customer_dfs[customer_name] = df

    # Create Excel file with the current date
    current_date = datetime.now().strftime('%Y-%m-%d')
    excel_file_path = f'{current_date}_New.xlsx'

    with pd.ExcelWriter(excel_file_path, engine='openpyxl') as writer:
        for customer_name, df in customer_dfs.items():
            df.to_excel(writer, sheet_name=customer_name, index=False, startrow=2)
    


    # Open the saved file with openpyxl to apply header styling
    wb = load_workbook(excel_file_path)

    # Define header names and merged cell ranges (as per your design)
    headers = [
        ("Customer", "A1:A2"),
        ("Facility", "B1:B2"),
        ("OWM7111", "C1:C2"),
        ("OWM0131", "D1:D2"),
        ("FXA", "E1:E2"),
        ("Common Area Camera", "F1:F2"),
        ("Thermostats", "G1:H1"),
        ("Centralize Temp & Humidity Sensor", "I1:I2"),
        ("Centralize Water Sensor", "J1:J2"),
        ("Centralize Door Sensor", "K1:K2"),
        ("Centralize Motion Sensor", "L1:L2"),
        ("R-Pi", "M1:M2"),
        ("Total", "N1:N2"),
    ]

    # Second row of information
    second_row = [
        "", "", "", "", "", "", 
        "ST898ZB" , "3157100", "3310-G", 
        "3315-G", "3323-G", "3328-G", "linuxevb",""
    ]

    # Apply styles and merge cells for each sheet
    header_font = Font(b=True, color="FFFFFF",size=22)
    header_fill = PatternFill("solid", fgColor="4F4F4F")
    header_alignment = Alignment(horizontal="center", vertical="center")
    border_style = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )

    for sheet_name in customer_dfs.keys():
        ws = wb[sheet_name]
        ws.merge_cells("G1:H1")
        # Write the first row of headers
        for header, cell_range in headers:
            start_cell, end_cell = cell_range.split(":")

            # Get the start cell for setting the value and applying the style
            cell = ws[start_cell]
            
            # Set the value and apply styles
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border_style

            
        # Write the second row of information
        for col_num, value in enumerate(second_row, start=1):
            cell = ws.cell(row=2, column=col_num)
            cell.value = value
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border_style
        # Adjust column widths based on the maximum length of the values in each column
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)  # Get the column letter
            
            # Find the maximum length of the string in the column
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            # Adjust the column width; add a little extra to prevent cutting off text
            adjusted_width = max_length + 8
            ws.column_dimensions[col_letter].width = adjusted_width
            
        ws.delete_rows(3)
        # Apply the Nunito font to all cells in the worksheet
        nunito_font = Font(name="Nunito")
        for row in ws.iter_rows():
            for cell in row:
                cell.font = nunito_font

        # Merge the last two rows
        last_row = ws.max_row
        last_column = ws.max_column
        second_last_row = last_row - 1
        
    
        for col_num in range(2, last_column + 1):  # Looping from column B to the last column
            cell = ws.cell(row=last_row, column=col_num)
            cell.font = Font(b=True,size=12,name="Nunito")
            
        # Define the style with no borders
        no_border = Border(left=Side(border_style=None),
                        right=Side(border_style=None),
                        top=Side(border_style=None),
                        bottom=Side(border_style=None))

        # Apply no border style starting from the 3rd row
        for row in ws.iter_rows(min_row=second_last_row,max_row=last_row):
            for cell in row:
                cell.border = Border(None,None,None,None)
        
        
                # Apply the white font style to the first two rows
        for row in ws.iter_rows(min_row=1, max_row=2):
            for cell in row:
                cell.font = Font(b=True,color='FFFFFF',size=12,name='Nunito')
    # Save the styled workbook
    wb.save(excel_file_path)
    wb.close()

    print(f"Header design with second row has been applied and saved to {excel_file_path}")
if __name__ == "__main__":
    process_data()
